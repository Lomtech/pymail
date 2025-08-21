# -*- coding: utf-8 -*-
"""
mailer.py – Excel -> Outlook (Windows) Mailer

Funktionen:
- Liest Kundenliste.xlsx (oder per -excel) und optional bestimmtes Blatt (-sheet).
- Baut korrekte Briefanrede (Herr/Frau/Titel).
- Rendert HTML mit Jinja2-Template (Standard: mail_template.html).
- Hängt auf Wunsch Outlook-Signatur an (-sig auto | none | <Name>).
- Versendet automatisch über Outlook (COM), optional -display zur Sichtprüfung.
- Optional-Spalten je Zeile:
    Betreff     : überschreibt globalen Betreff
    CC, BCC     : Kommagetrennt
    AnhangPfad  : Ein oder mehrere Pfade, getrennt durch ; oder ,
- Dry-Run mit -dry für Vorschau ohne Versand.
"""

import argparse
import os
import re
import sys
import time
import pathlib
from typing import Dict, List, Optional


from openpyxl import load_workbook
import win32com.client as win32
from jinja2 import Template


# ------------------------ Pfad-Utilities (EXE/Skript-sicher) ------------------------

def get_base_dir() -> pathlib.Path:
    """
    Ermittelt den Basisordner:
    - Bei PyInstaller --onefile: Ordner der ausgeführten EXE
    - Beim Skript: Ordner der .py-Datei
    """
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return pathlib.Path(sys.executable).parent
    return pathlib.Path(__file__).parent.resolve()


BASE_DIR = get_base_dir()


def resolve_path(p: Optional[str], default_name: str) -> pathlib.Path:
    """
    Pfad-Auflösung:
    1) Wenn ein Parameter p übergeben wurde:
       - absolut verwenden oder relativ zum AKTUELLEN Arbeitsverzeichnis (CWD)
    2) Wenn kein p:
       - Datei im BASE_DIR (Ordner der EXE/des Skripts) verwenden
    """
    if p:
        pp = pathlib.Path(p)
        return pp if pp.is_absolute() else (pathlib.Path.cwd() / pp)
    return BASE_DIR / default_name


# ------------------------ Hilfsfunktionen ------------------------

def normalize(s: Optional[str]) -> str:
    return (s or "").strip()


def read_subject(arg: Optional[str]) -> str:
    """Wenn arg eine Datei ist, deren Inhalt verwenden; sonst Text selbst."""
    if not arg:
        return ""
    p = pathlib.Path(arg)
    if p.exists() and p.is_file():
        return p.read_text(encoding="utf-8").strip()
    return arg.strip()


def build_anrede_brief(anrede: str, titel: str, vorname: str, nachname: str) -> str:
    a = normalize(anrede).lower()
    t = normalize(titel)
    ln = normalize(nachname)
    fn = " ".join(x for x in [normalize(vorname), normalize(nachname)] if x).strip()

    def pick(fb: str) -> str:
        if not t:
            return fb
        lower = t.lower()
        if lower.startswith("herr") or lower.startswith("frau"):
            return t  # Titel enthält bereits Anrede-Schlüsselwort → nicht doppeln
        return f"{fb} {t}"

    if a == "herr":
        return f"Sehr geehrter {pick('Herr')} {ln or fn}".strip()
    if a == "frau":
        return f"Sehr geehrte {pick('Frau')} {ln or fn}".strip()
    return f"Guten Tag {fn}" if fn else "Guten Tag"


def load_contacts(xlsx_path: str, sheet: Optional[str]) -> List[Dict[str, str]]:
    """Liest Excel; normalisiert Header zu lowercase; liefert nur Zeilen mit 'email'."""
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet:
        if sheet not in wb.sheetnames:
            raise SystemExit(f"Tabellenblatt '{sheet}' nicht gefunden (vorhanden: {wb.sheetnames})")
        ws = wb[sheet]
    else:
        ws = wb[wb.sheetnames[0]]  # erstes Blatt

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    headers = [normalize(str(h)).lower() for h in rows[0]]
    data: List[Dict[str, str]] = []
    for r in rows[1:]:
        rec = {}
        for i, v in enumerate(r):
            key = headers[i] if i < len(headers) else f"spalte_{i+1}"
            rec[key] = normalize(str(v)) if v is not None else ""
        if rec.get("email", ""):
            data.append(rec)
    return data


def find_signature_dir() -> pathlib.Path:
    appdata = os.environ.get("APPDATA", "")
    return pathlib.Path(appdata) / "Microsoft" / "Signatures"


def read_signature(sig_name: str) -> str:
    p = find_signature_dir() / f"{sig_name}.htm"
    if p.exists():
        return p.read_text(encoding="utf-8", errors="ignore")
    raise FileNotFoundError(p)


def read_default_signature() -> str:
    sig_dir = find_signature_dir()
    if not sig_dir.exists():
        raise FileNotFoundError(sig_dir)
    candidates = sorted(sig_dir.glob("*.htm"), key=lambda f: f.stat().st_mtime, reverse=True)
    if not candidates:
        raise FileNotFoundError("Keine .htm-Signatur gefunden")
    return candidates[0].read_text(encoding="utf-8", errors="ignore")


def render_html(template_path: str, context: Dict[str, str], signature_html: str = "") -> str:
    tpl = Template(pathlib.Path(template_path).read_text(encoding="utf-8"))
    html = tpl.render(**context)
    if signature_html:
        if "<!--SIGNATURE-->" in html or "<!--signature-->" in html:
            html = re.sub(r"<!--signature-->|<!--SIGNATURE-->", signature_html, html, flags=re.I)
        else:
            html += "<br>" + signature_html
    return html


def parse_list_field(val: str) -> List[str]:
    """Teilt CC/BCC/Anhänge über Komma oder Semikolon; trimmt; filtert Leere."""
    if not val:
        return []
    parts = re.split(r"[;,]", val)
    return [p.strip() for p in parts if p.strip()]


def add_attachments(mail, field_value: str):
    """Fügt Anhänge aus AnhangPfad hinzu (mehrere durch , oder ; getrennt)."""
    paths = parse_list_field(field_value)
    for p in paths:
        fp = pathlib.Path(p)
        if fp.exists() and fp.is_file():
            mail.Attachments.Add(str(fp))
        else:
            print(f"[WARN] Anhang nicht gefunden: {p}", file=sys.stderr)


def send_outlook(
    to_addr: str,
    subject: str,
    html: str,
    cc: str = "",
    bcc: str = "",
    attachments: str = "",
    display: bool = False,
):
    outlook = win32.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)  # olMailItem
    mail.To = to_addr
    if cc:
        mail.CC = cc
    if bcc:
        mail.BCC = bcc
    mail.Subject = subject
    mail.HTMLBody = html

    if attachments:
        add_attachments(mail, attachments)

    if display:
        mail.Display(False)  # manuell prüfen
    else:
        mail.Send()


# ------------------------ Main ------------------------

def main():
    parser = argparse.ArgumentParser(description="Excel -> Outlook Mailer (Windows)")
    parser.add_argument(
        "-excel",
        default="Kundenliste.xlsx",
        help="Pfad zur Excel-Datei (.xlsx), Standard: Kundenliste.xlsx",
    )
    parser.add_argument(
        "-sheet",
        default=None,
        help="Tabellenblatt-Name; Standard: erstes Blatt",
    )
    parser.add_argument(
        "-subject",
        default=None,
        help="Betreff-Text oder Pfad zu Textdatei; pro Zeile 'Betreff' überschreibt das",
    )
    parser.add_argument(
        "-template",
        default="mail_template.html",
        help="Pfad zur HTML-Template-Datei (Jinja2), Standard: mail_template.html",
    )
    parser.add_argument(
        "-sig",
        default="auto",
        help="Signatur: 'auto', 'none' oder exakter Name der Outlook-Signatur",
    )
    parser.add_argument(
        "-dry", action="store_true", help="Trockenlauf: keine E-Mails senden, nur Ausgabe"
    )
    parser.add_argument(
        "-display",
        action="store_true",
        help="E-Mails vor dem Senden anzeigen (statt sofort senden)",
    )
    parser.add_argument(
        "-throttle",
        type=float,
        default=0.3,
        help="Pause zwischen Mails (Sekunden), Standard: 0.3",
    )
    args = parser.parse_args()

    # Pfade robust auflösen (CWD vs. EXE/Skript-Ordner)
    excel_path = resolve_path(args.excel, "Kundenliste.xlsx")
    template_path = resolve_path(args.template, "mail_template.html")

    # Debug-Ausgaben (helfen bei 'Datei nicht gefunden')
    print(f"[DEBUG] BASE_DIR={BASE_DIR}")
    print(f"[DEBUG] CWD={pathlib.Path.cwd()}")
    print(f"[DEBUG] Excel={excel_path.resolve()}")
    print(f"[DEBUG] Template={template_path.resolve()}")

    if not excel_path.exists():
        raise SystemExit(f"Excel-Datei nicht gefunden: {excel_path.resolve()}")
    if not template_path.exists():
        raise SystemExit(f"Template-Datei nicht gefunden: {template_path.resolve()}")

    global_subject = read_subject(args.subject) if args.subject else ""

    # Kontakte laden
    contacts = load_contacts(str(excel_path), args.sheet)
    if not contacts:
        raise SystemExit("Keine Kontakte gefunden (mind. Spalte 'Email' und eine Datenzeile benötigt).")

    # Signatur bereitstellen
    signature_html = ""
    if args.sig.lower() == "auto":
        try:
            signature_html = read_default_signature()
        except Exception as e:
            print(f"[WARN] Standardsignatur nicht gefunden: {e}", file=sys.stderr)
    elif args.sig.lower() == "none":
        signature_html = ""
    else:
        try:
            signature_html = read_signature(args.sig)
        except Exception as e:
            print(f"[WARN] Signatur '{args.sig}' nicht gefunden: {e}", file=sys.stderr)

    print(
        f"[INFO] Kontakte: {len(contacts)} | Vorlage: {template_path.name} | "
        f"Signatur: {'auto' if args.sig.lower()=='auto' else args.sig} | Dry-Run: {args.dry}"
    )

    sent, failed = 0, 0
    for i, c in enumerate(contacts, start=1):
        email = c.get("email", "")
        if not email:
            continue

        # Betreff: pro Zeile 'betreff' > globaler Betreff > Fehler
        row_subject = c.get("betreff", "").strip()
        subject = row_subject or global_subject
        if not subject and args.dry is False:
            print(
                f"[SKIP] Kein Betreff für {email} (weder -subject noch Spalte 'Betreff').",
                file=sys.stderr,
            )
            failed += 1
            continue

        ctx = {
            "Email": email,
            "Anrede": c.get("anrede", ""),
            "Vorname": c.get("vorname", ""),
            "Nachname": c.get("nachname", ""),
            "Firma": c.get("firma", ""),
            "Titel": c.get("titel", ""),
        }
        ctx["AnredeBrief"] = build_anrede_brief(
            ctx["Anrede"], ctx["Titel"], ctx["Vorname"], ctx["Nachname"]
        )

        # Weitere Spalten ebenfalls bereitstellen (Originalschlüssel in lowercase)
        for k, v in c.items():
            if k.lower() not in {kk.lower() for kk in ctx.keys()}:
                ctx[k] = v

        # HTML rendern
        html = render_html(str(template_path), ctx, signature_html=signature_html)

        # CC/BCC/Anhänge
        cc_raw = c.get("cc", "")
        bcc_raw = c.get("bcc", "")
        attach_raw = c.get("anhangpfad", "") or c.get("anhang", "")

        if args.dry:
            print(
                f"\n--- DRY RUN #{i} ---"
                f"\nTO: {email}"
                f"\nSUBJECT: {subject or '(leer)'}"
                f"\nCC: {cc_raw}"
                f"\nBCC: {bcc_raw}"
                f"\nANHANG: {attach_raw}"
                f"\nHTML:\n{html}\n"
            )
            continue

        try:
            send_outlook(
                to_addr=email,
                subject=subject,
                html=html,
                cc=cc_raw,
                bcc=bcc_raw,
                attachments=attach_raw,
                display=args.display,
            )
            sent += 1
            print(f"[OK] {i}/{len(contacts)} → {email}")
        except Exception as e:
            failed += 1
            print(f"[ERR] {i}/{len(contacts)} → {email}: {e}", file=sys.stderr)

        # sanfte Drosselung
        time.sleep(args.throttle)

    if not args.dry:
        print(f"\nFertig. Gesendet: {sent}, Fehler: {failed}")


if __name__ == "__main__":
    main()
